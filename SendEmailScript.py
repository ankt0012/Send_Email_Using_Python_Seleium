from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd  # this library is used to read/extract the data from excel sheet
from openpyxl import load_workbook  # this library is used to edit the excel sheet
import os  # this library used to track system directory path to make this code dynamic so we will be able to run code in any system with just some instruction followed
import pyautogui # it is used to handle the gui while using chrome
import tkinter as tk
from tkinter import filedialog


email = ""
password = ""
current_directory = os.getcwd()
is_successful_login = False

# I use chrome options to disable blink feature that tracks automation code.
# I took this approach because google recent update, we are not able to login email through automation code
def start_webdriver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    s = Service(executable_path=rf"{current_directory}\chromedriver.exe")
    driver = webdriver.Chrome(service=s, options=chrome_options)
    driver.maximize_window()
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
                           Object.defineProperty(navigator, 'webdriver', {
                             get: () => undefined
                           })
                           """
    })
    return driver



# here i am reading excel sheet to extract data from it
df = pd.read_excel("credentials.xlsx")
receiver_email = df["Sender_email"].get(key=0)
subject = df["subject_for_email"].get(key=0)
file_name = df["File_name_with_file_type"].get(key=0)
body = df["body_for_email"].get(key=0)


# I am using class and methods to run actual code, in this class we are running two method
# 1st method login() - it takes email and password to login in gmail
# 2nd method send_email() - it takes receiver email where we want to send email.
class Login_gmail:

    def __init__(self):
        self.value=0

    def login(self, email, password,driver) :
        try :
            driver.get(
                "https://accounts.google.com/v3/signin/identifier?ifkv=AaSxoQzmRm1DeRuuqSwxSSN2aANYPTFZSjQU7J1EsMEFyvt4ihzgWcMfyiQk5KXZteBXmnG3a36a7Q&service=mail&flowName=GlifWebSignIn&flowEntry=ServiceLogin&dsh=S1275331952%3A1716101290562459&ddm=0")
            driver.implicitly_wait(10)
            wait = WebDriverWait(driver, 20)
            email_field = wait.until(EC.presence_of_element_located((By.ID, "identifierId")))
            email_field.send_keys(email)
            email_field.send_keys(Keys.RETURN)
            # Wait for the password field to appear
            time.sleep(5)  # Additional sleep in case network is slow
            password_field = wait.until(EC.presence_of_element_located((By.NAME, "Passwd")))
            password_field.send_keys(password)
            password_field.send_keys(Keys.RETURN)
            driver.implicitly_wait(20)
            time.sleep(10)
            print(driver.title)
            if  email in driver.title:
                is_successful_login = True
                print(True)
            else:
                is_successful_login = False
                print(False)
            time.sleep(5) # in case executable system has slow network
            return is_successful_login
        except Exception as e :
            print("having issue in login email \n",e)
            is_successful_login = False
            return is_successful_login

    def send_email(self, receiver_email,driver):
        try :

            action = ActionChains(driver)
            driver.find_element(By.XPATH,"//div[contains(text(), 'Compose')]").click()
            driver.find_element(By.ID, ":th").send_keys(receiver_email)
            subject_element = driver.find_element(By.ID, ":pv")
            subject_element.send_keys(f"{subject}")
            body_element = driver.find_element(By.ID, ":r5")
            action.move_to_element(body_element).perform()
            body_element.send_keys(f"{body}")
            attach_file = driver.find_element(By.ID, ":ri")
            action.move_to_element(attach_file).perform()
            attach_file.click()
            current_directory = os.getcwd()
            attachment = f"{file_name}"
            time.sleep(2)
            pyautogui.hotkey('alt', 'd')
            time.sleep(2)
            pyautogui.typewrite(f'{current_directory}')
            pyautogui.press('enter')
            time.sleep(2)
            pyautogui.hotkey('alt', 'n')
            pyautogui.typewrite(f'{attachment}')
            pyautogui.press('enter')
            time.sleep(2)
            send= driver.find_element(By.ID, ":pl")
            action.move_to_element(send).perform()
            send.click()
            time.sleep(2)
            driver.quit()
        except Exception as e :
            print("Have issue in sending email\n", e)
            driver.close()


my_class_instence = Login_gmail()

# Here i am itterate loop over the data available in excel sheet

def run() :
    wb = load_workbook('credentials.xlsx')
    ws = wb.active
    for index, value in df.iterrows() :

        if value["credential"] == "username" :
            email = value["actual_value"]
            cell1 =index+2
        elif value["credential"] == "password":
            password = value["actual_value"]
            cell2=index+2
        # Here i am implement condition where our login() method run after username & password field will be assigned, you ca
        if index % 2 == 1 :
            driver = start_webdriver()
            # print(f"this is before login{index}", f"{is_successful_login}")
            is_successful_loggedIn = my_class_instence.login(email, password, driver)

            ws.merge_cells(f'G{cell1}:G{cell2}')
            # Here i am Overwriting data in the worksheet
            if is_successful_loggedIn :
                cell= ws[f'G{cell1}']
                cell.value="Success"
                my_class_instence.send_email(receiver_email,driver)

            else :
                cell = ws[f'G{cell1}']
                cell.value = "Failure"
    wb.save('credentials.xlsx')

run()

